import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, Color, Fill
'''Este codigo pega o arquivo csv e transforma em uma versão excel ja filtrado com as informações necessarias'''

def cria_excel():
    print("Criando planilha excel")
    #Carrega o csv bruto do glpi
    chamados_df = pd.read_csv('glpi.csv', sep=';')

    #Cria um novo dataframe apenas com os valores necessarios
    novo_dataframe = chamados_df[["Atribuído para - Técnico", "ID", "Entidade"]]
    novo_dataframe.to_excel("provisorio.xlsx", sheet_name="Chamados Atribuidos")

    wb = load_workbook(filename = 'provisorio.xlsx')
    ws = wb['Chamados Atribuidos']

    ws['A1'].value = 'Quantidade'
    ws.cell(1,1).font = Font(bold= True)
    ws.cell(1,1).alignment = Alignment('center')

    i = 2
    while i <= ws.max_row:
        text = ws['D' + str(i)].value.split("-", 1)
        ws['D' + str(i)].value = text[1]
        i = i + 1

    ws['E1'].value = 'Status'
    ws.cell(1,5).font = Font(bold= True)
    ws.cell(1,5).alignment = Alignment('center')
    ws.auto_filter.ref = "A1:E1"

    i=2
    while i <= ws.max_row:
        ws.cell(i, 1).value += 1
        i+=1
    #Ajuste das dimensoes das colunas
    ws.column_dimensions["A"].width = 16.0
    ws.column_dimensions["B"].width = 33.0
    ws.column_dimensions["C"].width = 12.0
    ws.column_dimensions["D"].width = 55.0
    ws.column_dimensions["E"].width = 20.0
    
    #Colorindo as celulas do filtros
    cor = PatternFill(fill_type='solid',fgColor='00B0F0')

    i = 1
    while i < 6:
        ws.cell(1,i).fill = cor
        i+=1
    
    #Salvar a planilha
    wb.save("provisorio.xlsx")

   
    print("#Feito")


