import os
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection

#copia arquivo csv da pasta Downloads e deleta o arquivo apos isso
arquivocsv = "C:\\Downloads\glpi.csv"
destinocsv = "C:\\Users\Lucas\Documents\projetocsvglpi"
if os.path.isfile(arquivocsv):
    shutil.copy2(arquivocsv, destinocsv)
    os.remove(arquivocsv)
else:
    print("Erro %s arquivo não encontrado" % arquivocsv)

#Carrega o csv bruto do glpi
chamados_df = pd.read_csv('glpi.csv', sep=';')

#Transforma o dataframe chamados em excel
chamados_df.to_excel("arquivo.xlsx", sheet_name="Bruto", index=False)

#Cria um novo dataframe apenas com os valores necessarios
arquivoxls = chamados_df[["Atribuído para - Técnico", "ID", "Entidade"]]
#print(type(arquivoxls))

#Transforma esse dataframe em excel
arquivoxls.to_excel("provisorio.xlsx", sheet_name="Chamados Atribuidos")


#Carrega e trata esse Excel novo
wb = load_workbook(filename = 'provisorio.xlsx')
ws1 = wb['Chamados Atribuidos']


ws1['A1'].value = 'Quantidade'
ws1.cell(1,1).font = Font(bold= True)
ws1.cell(1,1).alignment = Alignment('center')


i = 2
while i <= ws1.max_row:
    text = ws1['D' + str(i)].value.split("-")
    ws1['D' + str(i)].value = text[1]
    i = i + 1

ws1['E1'].value = 'Status'

ws1.auto_filter.ref = "A1:E1"
#Salvar a planilha
wb.save("provisorio.xlsx")

#Cria um novo dataframe com esse excel
chamados_Atribuido_df = pd.read_excel("provisorio.xlsx", "Chamados Atribuidos")


#calcula o numero de chamados
qtd=len(chamados_Atribuido_df.value_counts(['ID']))

#Cria a lista de Analistas com chamados
usuarios = len(chamados_Atribuido_df.value_counts(['Atribuído para - Técnico']))
print('O numero de analistas com chamados: %d\n\n'%usuarios)
ListaAnalistas = chamados_Atribuido_df['Atribuído para - Técnico'].unique()
ListaAnalistas = list(ListaAnalistas)
#del(ListaAnalistas[0])
print(ListaAnalistas)


#Transforma em txt e organiza os valores no arquivo de texto
with open('chamados.txt', 'w') as f:
    pass

with open('chamados.txt', 'a') as f:
    f.write('TOTAL FIELD SP = %d\n'%qtd)
    f.write('%d analistas com chamados\n'%usuarios)
    f.write('----------------------------\n')

#Cria o texto para ser enviado para o whatsUp
for nome in ListaAnalistas:
    with open('chamados.txt', 'a') as f:
        f.writelines('*%s*\n'%nome)
    analista = chamados_Atribuido_df.loc[chamados_df['Atribuído para - Técnico'] == nome, ['ID','Entidade']]
    analista.to_csv("chamados.txt", header=None, index=None, sep='-', mode='a')
    with open('chamados.txt', 'a') as f:
        f.writelines('\n')
    

#with open('chamados.txt', 'w') as t:
analista = chamados_Atribuido_df.loc[chamados_df['Atribuído para - Técnico'] == 'Vinicius Oliveira', ['Atribuído para - Técnico','ID','Entidade']]


#Apaga o arquivo glpi.csv apos fazer os tratamentos
glpicsv = 'glpi.csv'
if os.path.isfile(glpicsv):
    os.remove(glpicsv)
    pass
else:
    print("Erro %s arquivo não encontrado" % glpicsv)
