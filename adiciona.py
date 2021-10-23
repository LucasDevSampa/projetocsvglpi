from openpyxl import load_workbook


wb = load_workbook(filename = 'Principal.xlsx')
wb1 = load_workbook(filename = 'provisorio.xlsx')

ws = wb1['Chamados Atribuidos']

linhas = ws.max_row
colunas = ws.max_column

print("%d Linhas"%linhas)
print('%d Colunas'%colunas)



if ws.cell(24,1).value == None:
    print("Nada escrito\n\n\n")

print(type(ws['A1':'E1']))

for row in ws.iter_rows(min_row=2, max_col=1, max_row=linhas):
    for cell in row:
        if cell.value == None:
            break
        print(cell.value)

print(wb1.sheetnames)