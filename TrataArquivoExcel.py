from re import A
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import numbers


def TratarArquivo(arquivo):
    #Carrega o arquivo Excel
    wb = load_workbook(filename = arquivo)
    ws1 = wb['Chamados Atribuidos']

    
    
    ws1['A1'].value = 'Quantidade'
    ws1.cell(1,1).font = Font(bold= True)
    ws1.cell(1,1).alignment = Alignment('center')
    print(ws1.max_row)

    #Trecho que trata a Coluna entidade
    i = 2
    while i <= ws1.max_row:
        text = ws1['D' + str(i)].value.split("-")
        ws1['D' + str(i)].value = text[1]
        i = i + 1

    #Trecho que trata a numeração da quantidade
    i = 2
    while i <= ws1.max_row:
        ws1['A' + str(i)].value += 1
        i = i + 1
   

    #Trecho que cria um Filtro
    ws1.auto_filter.ref = "A1:D1"

    #Criar as abas dos analistas Fields
    for cell in ws1['B']:
        wb.create_sheet(cell.value)
    

    

    

    #Salvar A planilha
    wb.save("FieldSP.xlsx")


    



   


